from io import BytesIO
from time import time
from datetime import datetime as dt
from random import Random
from colorsys import hls_to_rgb
from fastapi import APIRouter, Request, Response
from fastapi.responses import RedirectResponse, StreamingResponse, HTMLResponse
from openpyxl import Workbook, styles as wstyles
from playwright.async_api import async_playwright as pw
from ...models.user import User, Access
from ...models.zis_payment import Payment, PaymentCategory, PaymentUnit
from ...helpers.str_to_bool import str_to_bool
from ..dependencies import auth, NoAuthToken, UserSessionNotFound
from ..render import env as jenv, render
from ...helpers.datetime import days, months

PAYMENT_QUERY_NON_FILTER_PARAMS = {
    'include_deleted': str_to_bool,
    'include_undeleted': str_to_bool,
    'sort': str,
    'limit': int,
    'offset': int,
}
parse_query_params = lambda params: {'filters': {key.replace('-', '_'): value for key, value in dict(params).items() if key.replace('-', '_') not in PAYMENT_QUERY_NON_FILTER_PARAMS}, **{key.replace('-', '_'): PAYMENT_QUERY_NON_FILTER_PARAMS[key.replace('-', '_')](value.replace('-', '_')) for key, value in dict(params).items() if key.replace('-', '_') in PAYMENT_QUERY_NON_FILTER_PARAMS}}

router = APIRouter()
payments = APIRouter()

@router.get('')
async def root(): return RedirectResponse(url=f'/{__name__.split(".")[-1]}/payments', status_code=302)

@payments.get('')
async def payments_index(request: Request):
    try:
        user = await auth(request)
        await user.require_access(Access.ZIS_PAYMENT_READ)
    except (NoAuthToken, UserSessionNotFound): return RedirectResponse(url='/login', status_code=302)
    except Access.AccessDenied: return RedirectResponse(url='/home', status_code=302)
    return await render('pages/zis/payments', {'user': user, 'payments': [(await (await payment.latest).to_dict) for payment in await Payment.query(**parse_query_params(request.query_params))]}, expose='payments')

@payments.get('/filter')
async def payments_filter(request: Request):
    try:
        user = await auth(request)
        await user.require_access(Access.ZIS_PAYMENT_READ)
    except (NoAuthToken, UserSessionNotFound): return RedirectResponse(url='/login', status_code=302)
    except Access.AccessDenied: return RedirectResponse(url='/home', status_code=302)
    return await render('pages/zis/payments/filter', {'user': user})

@payments.get('/export-xlsx')
async def payments_xlsx(request: Request):
    try:
        user = await auth(request)
        await user.require_access(Access.ZIS_PAYMENT_READ)
    except (NoAuthToken, UserSessionNotFound): return RedirectResponse(url='/login', status_code=302)
    except Access.AccessDenied: return RedirectResponse(url='/home', status_code=302)
    wb = Workbook()
    ws = wb.active
    if ws is None: raise RuntimeError('Failed to get active sheet of workbook object (using openpyxl).')
    ws.title = 'Pembayaran Zakat'
    ws.append((
        'Nomor Zakat',
        'Nomor Transaksi',
        'Nama',
        'Alamat',
        'Kategori',
        None,
        None,
        None,
        'Beras',
        None,
        'Admin',
        'Edit',
        'UUID',
    ))
    for start, end in ((5, 8), (9, 10)): ws.merge_cells(
        start_row=ws.max_row,
        start_column=start,
        end_row=ws.max_row,
        end_column=end,
    )
    ws.append((
        *(None,) * 4,
        'Fitrah',
        'Maal',
        'Fidyah',
        'Infaq',
        'Kilogram',
        'Liter',
    ))
    for col in (*range(1, 5), *range(11, 14)): ws.merge_cells(
        start_row=ws.max_row - 1,
        start_column=col,
        end_row=ws.max_row,
        end_column=col,
    )
    counter = 0
    users = {}
    for payment in (await Payment.query(**parse_query_params(request.query_params))):
        counter += 1
        payment = await (await payment.latest).to_dict
        created_by = users[username] if (username := payment['created_by']) in users else (await User.get(username=username))
        if username not in users: users[username] = created_by
        updated_by = users[username] if (username := payment['updated_by']) in users else (await User.get(username=username))
        if username not in users: users[username] = updated_by
        ws.append(())
        ws.merge_cells(
            start_row=ws.max_row+1,
            start_column=1,
            end_row=ws.max_row+1,
            end_column=14,
        )
        ws.append((*(f'\'x' if isinstance(x, str) and x.startswith('=') else x for x in (
            counter,
            1,
            payment['lines'][0]['payer_name'],
            payment['payer_address'],
            (payment['lines'][0]['amount'] if payment['lines'][0]['unit'] == 'rupiah' else 0) if payment['lines'][0]['category'] == 'zakat fitrah' else None,
            (payment['lines'][0]['amount'] if payment['lines'][0]['unit'] == 'rupiah' else 0) if payment['lines'][0]['category'] == 'zakat maal' else None,
            (payment['lines'][0]['amount'] if payment['lines'][0]['unit'] == 'rupiah' else 0) if payment['lines'][0]['category'] == 'fidyah' else None,
            (payment['lines'][0]['amount'] if payment['lines'][0]['unit'] == 'rupiah' else 0) if payment['lines'][0]['category'] == 'infaq' else None,
            payment['lines'][0]['amount'] if payment['lines'][0]['unit'] == 'kilogram beras' else None,
            payment['lines'][0]['amount'] if payment['lines'][0]['unit'] == 'liter beras' else None,
            f'{created_by.name} ({created_by.username})' if created_by is not None else '[Gagal mendapatkan informasi user]',
            f'{updated_by.name} ({updated_by.username})' if updated_by is not None else '[Gagal mendapatkan informasi user]',
            payment['payment'],
        )),))
        for col in range(5, 9): ws.cell(
            row=ws.max_row,
            column=col,
        ).number_format = '[$Rp-421]#,##0'
        for col in (
            {'col': 11, 'username': created_by.username if created_by is not None else ''},
            {'col': 12, 'username': updated_by.username if updated_by is not None else ''},
        ): ws.cell(
            row=ws.max_row,
            column=col['col'],
        ).fill = wstyles.PatternFill(
            start_color=(lambda x: f'{round((c := hls_to_rgb(Random(x).randint(0, 255) / 255, 0.5, 1))[0] * 255):02X}{round(c[1] * 255):02X}{round(c[2] * 255):02X}')(col['username']),
            fill_type='solid',
        )
        line_counter = 1
        for line in payment['lines'][1:]:
            line_counter += 1
            ws.append((*(f'\'x' if isinstance(x, str) and x.startswith('=') else x for x in (
                None,
                line_counter,
                line['payer_name'],
                None,
                (line['amount'] if line['unit'] == 'rupiah' else 0) if line['category'] == 'zakat fitrah' else None,
                (line['amount'] if line['unit'] == 'rupiah' else 0) if line['category'] == 'zakat maal' else None,
                (line['amount'] if line['unit'] == 'rupiah' else 0) if line['category'] == 'fidyah' else None,
                (line['amount'] if line['unit'] == 'rupiah' else 0) if line['category'] == 'infaq' else None,
                line['amount'] if line['unit'] == 'kilogram beras' else None,
                line['amount'] if line['unit'] == 'liter beras' else None,
            )),))
            for col in range(5, 9): ws.cell(
                row=ws.max_row,
                column=col,
            ).number_format = '[$Rp-421]#,##0'
        for col in (1, 4, *range(11, 14)):
            if len(payment['lines']) > 1: ws.merge_cells(
                start_row=ws.max_row - len(payment['lines']) + 1,
                start_column=col,
                end_row=ws.max_row,
                end_column=col,
            )
        for col in 'ABCDEFGHIJKLM': ws.column_dimensions[col].auto_size = True
    wb.save(buffer := BytesIO())
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'{"attachment" if "download" in dict(request.query_params).keys() else "inline"}; filename=Export_Pembayaran_ZIS_{int(time())}.xlsx'}
    )

@payments.get('/new')
async def payment_new(request: Request, response: Response):
    try:
        user = await auth(request)
        await user.require_access(Access.ZIS_PAYMENT_READ, Access.ZIS_PAYMENT_WRITE)
    except (NoAuthToken, UserSessionNotFound): return RedirectResponse(url='/login', status_code=302)
    except Access.AccessDenied: return RedirectResponse(url=f'/payments' if user.has_access(Access.ZIS_PAYMENT_READ) else '/home', status_code=302)
    return await render('pages/zis/payments/new', {'user': user, 'categories': (*({'value': category, 'name': category.capitalize()} for id, category in (await PaymentCategory.get_all()).items()),), 'units': (*({'value': unit, 'name': unit.capitalize()} for id, unit in (await PaymentUnit.get_all()).items()),)})

@payments.get('/{uuid}')
async def payment(request: Request, response: Response, uuid: str):
    try:
        user = await auth(request)
        await user.require_access(Access.ZIS_PAYMENT_READ)
    except (NoAuthToken, UserSessionNotFound): return RedirectResponse(url='/login', status_code=302)
    except Access.AccessDenied: return RedirectResponse(url='/home', status_code=302)
    payment = await Payment.get(uuid=uuid)
    if payment is None: return await render('pages/error', {'code': '404', 'error': 'Pembayaran Tidak Ditemukan', 'user': user}, status_code=404)
    return await render('pages/zis/payments/show', {'user': user, 'payment': await (await payment.latest).to_dict}, expose='payment')

@payments.get('/{uuid}/edit')
async def payment_edit(request: Request, response: Response, uuid: str):
    try:
        user = await auth(request)
        await user.require_access(Access.ZIS_PAYMENT_READ, Access.ZIS_PAYMENT_WRITE)
    except (NoAuthToken, UserSessionNotFound): return RedirectResponse(url='/login', status_code=302)
    except Access.AccessDenied: return RedirectResponse(url=f'/payments/{uuid}' if user.has_access(Access.ZIS_PAYMENT_READ) else '/home', status_code=302)
    payment = await Payment.get(uuid=uuid)
    if payment is None: return await render('pages/error', {'code': '404', 'error': 'Pembayaran Tidak Ditemukan', 'user': user}, status_code=404)
    return await render('pages/zis/payments/edit', {'user': user, 'payment': await (await payment.latest).to_dict, 'categories': (*({'value': category, 'name': category.capitalize()} for id, category in (await PaymentCategory.get_all()).items()),), 'units': (*({'value': unit, 'name': unit.capitalize()} for id, unit in (await PaymentUnit.get_all()).items()),)}, expose='payment')

@payments.get('/{uuid}/receipt')
async def payment_receipt(request: Request, response: Response, uuid: str):
    query_params = dict(request.query_params)
    try:
        user = await auth(request)
        await user.require_access(Access.ZIS_PAYMENT_READ)
    except (NoAuthToken, UserSessionNotFound): return RedirectResponse(url='/login', status_code=302)
    except Access.AccessDenied: return RedirectResponse(url='/home', status_code=302)
    payment = await Payment.get(uuid=uuid)
    if payment is None: return await render('pages/error', {'code': '404', 'error': 'Pembayaran Tidak Ditemukan', 'user': user}, status_code=404)
    payment = await (await payment.latest).to_dict
    ca = dt.fromtimestamp(payment['created_at'])
    payment['created_at'] = f'{days[(ca.weekday() + 1) % 7]}, {ca.day} {months[ca.month - 1]} {ca.year}'
    payment['created_by'] = await User.get(username=payment['created_by'])
    if payment['created_by'] is None: raise RuntimeError('Failed to fetch user from payment.created_by')
    payment['created_by'] = payment['created_by'].name
    html = jenv.get_template('pdf/zis_payment_receipt.html').render(payment, format_number=lambda x: f'{x:,}')
    if 'html' in query_params: return HTMLResponse(html)
    async with pw() as p:
        browser = await p.chromium.launch(headless=True)
        try:
            page = await browser.new_page()
            try:
                await page.set_content(html)
                pdf = await page.pdf(format=query_params['format'] if 'format' in query_params else 'A5', print_background=True)
            finally: await page.close()
        finally: await browser.close()
    return StreamingResponse(
        BytesIO(pdf),
        media_type='application/pdf',
        headers={'Content-Disposition': f'{"attachment" if "download" in query_params else "inline"}; filename={payment["payment"]}.pdf'}
    )

router.include_router(payments, prefix='/payments')