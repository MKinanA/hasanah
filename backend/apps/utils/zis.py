from datetime import datetime as dt, date
from time import time
from colorsys import hls_to_rgb
from random import Random
from playwright.async_api import async_playwright as pw
from pdfkit import from_string as pdfkit_from_string
from openpyxl import Workbook
from openpyxl import styles as wstyles
from openpyxl.worksheet.worksheet import Worksheet
from ...models.user import User
from ...models.zis_payment import Payment, PaymentVersion
from ...helpers import sql_commands as sql
from ...helpers.datetime import days, months
from ..render import env as jenv
from ...helpers.log import log
from ...helpers.db_connect import db_connect
from ...helpers.datetime import days, months

DEBUG_FORCE_PDFKIT = False

async def generate_receipt(payment: 'Payment | PaymentVersion | dict', format: 'str | None' = None, html: 'bool' = False) -> 'bytes | str':
    if isinstance(payment, Payment): payment = await payment.latest
    if isinstance(payment, PaymentVersion): payment = await payment.to_dict
    payment['units'] = []
    for line in payment['lines']:
        if not line['unit'] in payment['units']: payment['units'].append(line['unit'])
    ca = dt.fromtimestamp(payment['created_at'])
    payment['created_at'] = f'{days[(ca.weekday() + 1) % 7]}, {ca.day} {months[ca.month - 1]} {ca.year}'
    payment['created_by'] = await User.get(username=payment['created_by'])
    if payment['created_by'] is None: raise RuntimeError('Failed to fetch user from payment.created_by')
    payment['created_by'] = payment['created_by'].name
    payment['updated_by'] = await User.get(username=payment['updated_by'])
    if payment['updated_by'] is None: raise RuntimeError('Failed to fetch user from payment.updated_by')
    payment['updated_by'] = payment['updated_by'].name
    payment['totals'] = {}
    for line in payment['lines']:
        if line['unit'] not in payment['totals']: payment['totals'][line['unit']] = 0
        payment['totals'][line['unit']] += line['amount']
    html_render = jenv.get_template('pdf/zis_payment_receipt.html').render(payment, format_number=lambda x: f'{x:,}')
    if html: return html_render
    try:
        if DEBUG_FORCE_PDFKIT: raise NotImplementedError('Forced pdfkit for testing purposes.')
        async with pw() as p:
            browser = await p.chromium.launch(headless=True)
            try:
                page = await browser.new_page()
                try:
                    await page.set_content(html_render)
                    pdf = await page.pdf(format=format if isinstance(format, str) else 'A5', print_background=True)
                finally: await page.close()
            finally: await browser.close()
    except NotImplementedError as e:
        print(log(__name__, f'Playwright PDF generation failed, falling back to pdfkit. Error: {type(e).__name__}({e})'))
        html_render = jenv.get_template('pdf/zis_payment_receipt_wkhtmltopdf.html').render(payment, format_number=lambda x: f'{x:,}')
        pdf = pdfkit_from_string(html_render, False, options={
            'page-size': format if isinstance(format, str) else 'A5',
            'print-media-type': '',
            'disable-smart-shrinking': '',
        })
        if not isinstance(pdf, bytes): raise RuntimeError('pdfkit did not return bytes.')
    return pdf

async def generate_report_custom_new(time_range_seconds: 'int | None' = None) -> Workbook:
    payments = (*(await Payment.query(filters={
        **({'first_created_in_timespan': time_range_seconds,} if time_range_seconds is not None else {}),
    }, sort='first_created')), None)
    wb = Workbook()
    first_ws = wb.active
    if first_ws is None: raise RuntimeError('Failed to get active sheet of workbook object (using openpyxl).')

    prev_date = None
    counter = 0
    day_counter = 0
    day_sum = {
        'jumlah pembayaran': 0,
        'jumlah jiwa fitrah': 0,
        'zakat fitrah': 0.0,
        'zakat maal': 0.0,
        'fidyah': 0.0,
        'infaq': 0.0,
        'beras kilogram': 0.0,
        'beras liter': 0.0,
    }
    days_sum: 'dict[str, dict]' = {}
    for payment in payments:
        if payment is not None:
            payment_lv, payment_fv = await payment.latest, await payment.version()
            if payment_lv is None: raise RuntimeError('Failed to fetch last version of a payment.')
            if payment_fv is None: raise RuntimeError('Failed to fetch first version of a payment.')
            payment_dt = dt.fromtimestamp(payment_fv.created_at or 0)
            payment_date = payment_dt.date()

        if prev_date is None or prev_date != payment_date or payment is None:
            if prev_date is not None or payment is None:
                if ws:
                    ws.append((
                        None,
                        None,
                        None,
                        'Total',
                        day_sum['jumlah jiwa fitrah'],
                        day_sum['zakat fitrah'],
                        day_sum['zakat maal'],
                        day_sum['fidyah'],
                        day_sum['infaq'],
                        day_sum['beras kilogram'],
                        day_sum['beras liter'],
                    ))
                    for col in range(6, 10): ws.cell(
                        row=ws.max_row,
                        column=col,
                    ).number_format = '[$Rp-421]#,##0'
                days_sum[payment_date_str] = day_sum.copy()
                for key in day_sum.keys(): day_sum[key] = 0.0

                if payment is None: continue

            prev_date = payment_date
            day_counter = 0
            payment_date_str = f'{days[payment_date.weekday()]}, {payment_date.day} {months[payment_date.month - 1]} {payment_date.year}'
            ws: Worksheet = wb.create_sheet(payment_date_str)
            ws.append((payment_date_str,))
            ws.merge_cells(
                start_row=ws.max_row,
                start_column=1,
                end_row=ws.max_row,
                end_column=14,
            )
            ws.append((
                'No.',
                None,
                'Waktu',
                'Nama',
                'Jumlah Jiwa Fitrah',
                'Uang',
                None,
                None,
                None,
                'Beras',
                None,
                'Petugas',
                'Edit Terakhir',
                'UUID',
            ))
            for start, end in ((1, 2), (6, 9), (10, 11)): ws.merge_cells(
                start_row=ws.max_row,
                start_column=start,
                end_row=ws.max_row,
                end_column=end,
            )
            ws.append((
                None,
                None,
                None,
                None,
                None,
                'Fitrah',
                'Mall',
                'Fidyah',
                'Infaq',
                'KG',
                'L',
                None,
                None,
                None,
            ))
            for start, end in ((1, 2), (3, 3), (4, 4), (5, 5), (12, 12), (13, 13), (14, 14)): ws.merge_cells(
                start_row=ws.max_row - 1,
                start_column=start,
                end_row=ws.max_row,
                end_column=end,
            )

        counter += 1
        day_counter += 1
        payment_lines = await payment_lv.lines
        payment_time = payment_dt.time()

        day_sum['jumlah pembayaran'] += 1
        day_sum['jumlah jiwa fitrah'] += (line_jf_sum := len((*(line for line in payment_lines if line.category == 'zakat fitrah'),)))
        day_sum['zakat fitrah'] += (line_mzf_sum := sum((*(line.amount for line in payment_lines if line.category == 'zakat fitrah' and line.unit == 'rupiah'),)))
        day_sum['zakat maal'] += (line_mzm_sum := sum((*(line.amount for line in payment_lines if line.category == 'zakat maal' and line.unit == 'rupiah'),)))
        day_sum['fidyah'] += (line_f_sum := sum((*(line.amount for line in payment_lines if line.category == 'fidyah' and line.unit == 'rupiah'),)))
        day_sum['infaq'] += (line_i_sum := sum((*(line.amount for line in payment_lines if line.category == 'infaq' and line.unit == 'rupiah'),)))
        day_sum['beras kilogram'] += (line_rkg_sum := sum((*(line.amount for line in payment_lines if line.unit == 'kilogram beras'),)))
        day_sum['beras liter'] += (line_rl_sum := sum((*(line.amount for line in payment_lines if line.unit == 'liter beras'),)))

        ws.append((*(f'\'x' if isinstance(x, str) and x.startswith('=') else x for x in (
            counter,
            day_counter,
            f'{payment_time.hour:02}:{payment_time.minute:02}:{payment_time.second:02}',
            payment_lv.payer_name,
            line_jf_sum or None,
            line_mzf_sum or None,
            line_mzm_sum or None,
            line_f_sum or None,
            line_i_sum or None,
            line_rkg_sum or None,
            line_rl_sum or None,
            payment_fv.created_by.name,
            payment_lv.created_by.name,
            payment.uuid,
        )),))
        for col in range(6, 10): ws.cell(
            row=ws.max_row,
            column=col,
        ).number_format = '[$Rp-421]#,##0'
        for col in (
            {'col': 12, 'username': payment_fv.created_by.name},
            {'col': 13, 'username': payment_lv.created_by.name},
        ): ws.cell(
            row=ws.max_row,
            column=col['col'],
        ).fill = wstyles.PatternFill(
            start_color=(lambda x: f'{round((c := hls_to_rgb(Random(x).randint(0, 255) / 255, 0.5, 1))[0] * 255):02X}{round(c[1] * 255):02X}{round(c[2] * 255):02X}')(col['username']),
            fill_type='solid',
        )

    ws = first_ws
    ws.title = 'Rekap'
    ws.append((f'{(*days_sum.keys(),)[0]} - {(*days_sum.keys(),)[-1]}',))
    ws.merge_cells(
        start_row=ws.max_row,
        start_column=1,
        end_row=ws.max_row,
        end_column=10,
    )
    ws.append((
        'No.',
        'Hari/Tanggal',
        'Total Pembayaran',
        'Total Jiwa Fitrah',
        'Total Uang',
        None,
        None,
        None,
        'Total Beras',
        None,
    ))
    for start, end in ((5, 8), (9, 10)): ws.merge_cells(
        start_row=ws.max_row,
        start_column=start,
        end_row=ws.max_row,
        end_column=end,
    )
    ws.append((
        'No.',
        'Hari/Tanggal',
        'Total Pembayaran',
        'Total Jiwa Fitrah',
        'Fitrah',
        'Mall',
        'Fidyah',
        'Infaq',
        'KG',
        'L',
    ))
    for start, end in ((1, 1), (2, 2), (3, 3), (4, 4)): ws.merge_cells(
        start_row=ws.max_row - 1,
        start_column=start,
        end_row=ws.max_row,
        end_column=end,
    )

    counter = 0
    final_sum = {key: 0.0 for key in day_sum.keys()}
    for day, day_sum in days_sum.items():
        counter += 1
        final_sum['jumlah pembayaran'] += (day_jp_sum := day_sum['jumlah pembayaran'])
        final_sum['jumlah jiwa fitrah'] += (day_jjf_sum := day_sum['jumlah jiwa fitrah'])
        final_sum['zakat fitrah'] += (day_zf_sum := day_sum['zakat fitrah'])
        final_sum['zakat maal'] += (day_zm_sum := day_sum['zakat maal'])
        final_sum['fidyah'] += (day_f_sum := day_sum['fidyah'])
        final_sum['infaq'] += (day_i_sum := day_sum['infaq'])
        final_sum['beras kilogram'] += (day_bk_sum := day_sum['beras kilogram'])
        final_sum['beras liter'] += (day_bl_sum := day_sum['beras liter'])
        ws.append((
            counter,
            day,
            day_jp_sum or None,
            day_jjf_sum or None,
            day_zf_sum or None,
            day_zm_sum or None,
            day_f_sum or None,
            day_i_sum or None,
            day_bk_sum or None,
            day_bl_sum or None,
        ))
        for col in range(5, 9): ws.cell(
            row=ws.max_row,
            column=col,
        ).number_format = '[$Rp-421]#,##0'

    ws.append((
        None,
        'Total',
        final_sum['jumlah pembayaran'],
        final_sum['jumlah jiwa fitrah'],
        final_sum['zakat fitrah'],
        final_sum['zakat maal'],
        final_sum['fidyah'],
        final_sum['infaq'],
        final_sum['beras kilogram'],
        final_sum['beras liter'],
    ))
    for col in range(5, 9): ws.cell(
        row=ws.max_row,
        column=col,
    ).number_format = '[$Rp-421]#,##0'

    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row: cell.border = wstyles.Border(
                left=wstyles.Side(style='thin'),
                right=wstyles.Side(style='thin'),
                top=wstyles.Side(style='thin'),
                bottom=wstyles.Side(style='thin')
            )
    return wb

async def change_pv_created_at(pv: PaymentVersion, to: int):
    if pv.id is None: raise ValueError('`pv` doesn\'t have an id.')
    async with db_connect() as conn:
        cursor = await conn.cursor()
        await cursor.execute(*sql.update('zis_payment_version', where={'id': pv.id}, set={'created_at': to}))
        await conn.commit()